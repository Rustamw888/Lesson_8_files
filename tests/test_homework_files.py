import csv
import os
import zipfile

from PyPDF2 import PdfReader
from openpyxl import load_workbook


def test_zip_files():
    homework_zip = zipfile.ZipFile('../resources/archive_files.zip', 'w')
    for folder, subfolders, files in os.walk('../resources'):

        for file in files:
            if file.endswith('.csv') or file.endswith('.pdf') or file.endswith('.txt') or file.endswith('.xlsx'):
                homework_zip.write(os.path.join(folder, file), os.path.relpath
                (os.path.join(folder, file), '../resources'), compress_type=zipfile.ZIP_DEFLATED)
    homework_zip.close()


def test_unpacked_files():
    fantasy_zip = zipfile.ZipFile('../resources/archive_files.zip')
    fantasy_zip.extractall('../tests/unpacked_files/')
    fantasy_zip.close()


def test_csv():
    with open('../tests/unpacked_files/csvSample.csv') as f:
        reader = csv.reader(f)
        headers = next(reader)
    assert 'Data_value' in str(headers)


def test_pdf():
    reader = PdfReader('../tests/unpacked_files/pdfSample.pdf')
    page = reader.pages[2]
    text = page.extract_text()
    assert 'Background' in text


def test_txt():
    reader = open('../tests/unpacked_files/txtSample.txt')
    assert reader.readline() == 'Quod equidem non reprehendo;\n'


def test_xlsx():
    workbook = load_workbook('../tests/unpacked_files/xlsxSample.xlsx')
    sheet = workbook.active
    result = sheet.cell(row=32, column=6).value
    assert 1875 == result
